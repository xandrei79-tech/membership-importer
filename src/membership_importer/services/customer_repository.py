"""Repository for loading customer records from the customer workbook."""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from ..models.customer import Customer
from ..models.member import Member
from ..models.payment_group import PaymentGroup

_REQUIRED_COLUMNS = (
    "customer_id",
    "customer_name",
    "group_name",
    "description",
    "monthly_amount",
    "currency",
    "mac",
    "full_name",
    "active",
)


class CustomerRepository:
    """Read customer configuration from the customers workbook."""

    def __init__(self, path: str | Path | None = None) -> None:
        workbook_path = Path(path) if path is not None else self._default_path()
        self._path = workbook_path
        self._customers: dict[str, Customer] = {}
        self._members_by_mac: dict[str, Member] = {}
        self._groups_by_name: dict[str, PaymentGroup] = {}
        self._load()

    @staticmethod
    def _default_path() -> Path:
        return Path(__file__).resolve().parents[3] / "config" / "customers.xlsx"

    def _load(self) -> None:
        if not self._path.exists():
            raise FileNotFoundError(f"Customer workbook not found: {self._path}")

        workbook = load_workbook(self._path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not rows:
            return

        headers = [self._normalize_header(value) for value in rows[0]]
        missing = [column for column in _REQUIRED_COLUMNS if column not in headers]
        if missing:
            joined = ", ".join(missing)
            raise ValueError(f"Missing required columns: {joined}")

        for row in rows[1:]:
            if row is None or not any(value is not None and str(value).strip() for value in row):
                continue

            values = dict(zip(headers, row, strict=False))
            customer_id = self._clean_text(values.get("customer_id"))
            if not customer_id:
                continue

            customer = self._customers.setdefault(
                customer_id,
                Customer(
                    customer_id=customer_id,
                    customer_name=self._clean_text(values.get("customer_name")),
                ),
            )
            customer.customer_name = customer.customer_name or self._clean_text(
                values.get("customer_name")
            )

            group_name = self._clean_text(values.get("group_name"))
            if group_name:
                group = self._groups_by_name.setdefault(
                    group_name,
                    PaymentGroup(
                        name=group_name,
                        description=self._clean_text(values.get("description")),
                        monthly_amount=self._decimal_value(
                            values.get("monthly_amount"), group_name
                        ),
                        currency=self._clean_text(values.get("currency")) or "EUR",
                    ),
                )
                customer.group = group
                if customer not in group.customers:
                    group.customers.append(customer)

            mac = self._clean_text(values.get("mac"))
            full_name = self._clean_text(values.get("full_name"))
            if mac:
                member = Member(
                    mac=mac,
                    full_name=full_name,
                    active=self._coerce_bool(values.get("active")),
                )
                existing_member = self._members_by_mac.get(mac)
                if existing_member is None:
                    self._members_by_mac[mac] = member
                else:
                    member = existing_member
                if member not in customer.members:
                    customer.members.append(member)

    def get_customer(self, customer_id: str) -> Customer | None:
        """Return the customer matching ``customer_id``."""
        return self._customers.get(self._clean_text(customer_id))

    def get_member(self, mac: str) -> Member | None:
        """Return the member matching ``mac``."""
        return self._members_by_mac.get(self._clean_text(mac))

    def get_group(self, customer_id: str) -> PaymentGroup | None:
        """Return the payment group for the customer matching ``customer_id``."""
        customer = self.get_customer(customer_id)
        return customer.group if customer is not None else None

    def list_groups(self) -> list[PaymentGroup]:
        """Return all unique payment groups."""
        return list(self._groups_by_name.values())

    @staticmethod
    def _normalize_header(value: object) -> str:
        return str(value).strip().lower().replace(" ", "_")

    @staticmethod
    def _clean_text(value: object) -> str:
        return str(value).strip() if value is not None else ""

    @staticmethod
    def _coerce_bool(value: object) -> bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        text = str(value).strip().lower()
        if text in {"true", "yes", "y", "1"}:
            return True
        if text in {"false", "no", "n", "0", ""}:
            return False
        return bool(text)

    @staticmethod
    def _decimal_value(value: object, field_name: str) -> Decimal:
        if value is None or value == "":
            return Decimal("0")
        try:
            return Decimal(str(value))
        except InvalidOperation as exc:  # pragma: no cover - defensive guard
            raise ValueError(f"Invalid value for {field_name}: {value!r}") from exc

    @property
    def path(self) -> Path:
        """Return the workbook path used by the repository."""
        return self._path
