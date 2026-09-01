"""Interfaces for reading and writing membership workbooks."""

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


@dataclass(frozen=True)
class WorkbookAnalysisResult:
    """Store a loaded workbook and its inspected worksheet metadata."""

    workbook: Workbook
    worksheet_names: tuple[str, ...]
    active_worksheet_name: str
    required_worksheets: dict[str, bool]


class ExcelManager:
    """Coordinate workbook access while preserving the domain boundary."""

    def load_workbook(self, path: Path) -> WorkbookAnalysisResult:
        """Load ``path`` and inspect its worksheet structure in memory."""
        workbook = load_workbook(path)
        worksheet_names = tuple(workbook.sheetnames)
        active_worksheet_name = workbook.active.title
        required_worksheets = {
            year: year in worksheet_names
            for year in ("2025", "2026", "2027")
        }
        return WorkbookAnalysisResult(
            workbook=workbook,
            worksheet_names=worksheet_names,
            active_worksheet_name=active_worksheet_name,
            required_worksheets=required_worksheets,
        )

    def save_workbook(self, path: Path) -> None:
        """Save the current workbook to ``path``."""
        raise NotImplementedError("Workbook management is not implemented yet")