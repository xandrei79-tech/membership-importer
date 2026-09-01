from pathlib import Path

from openpyxl import Workbook

from membership_importer.bank_importer import BankImporter
from membership_importer.models.payment import Payment
from membership_importer.services.import_service import ImportService


def test_import_payments_rejects_statements_without_an_implemented_importer(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "membership.xlsx"
    first_statement_path = tmp_path / "statement-one.csv"
    second_statement_path = tmp_path / "statement-two.pdf"

    workbook = Workbook()
    workbook.active.title = "2025"
    workbook.save(workbook_path)
    first_statement_path.write_bytes(b"first statement")
    second_statement_path.write_bytes(b"second statement")

    try:
        ImportService().import_payments(
            workbook_path,
            (first_statement_path, second_statement_path),
        )
    except NotImplementedError as error:
        assert str(error) == "No bank importer is implemented for 'statement-one.csv'."
    else:
        raise AssertionError("An unimplemented bank importer should be rejected")


def test_import_payments_collects_payments_from_detected_importers(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "membership.xlsx"
    first_statement_path = tmp_path / "statement-one.csv"
    second_statement_path = tmp_path / "statement-two.pdf"

    workbook = Workbook()
    workbook.save(workbook_path)
    first_statement_path.write_bytes(b"first statement")
    second_statement_path.write_bytes(b"second statement")

    first_payment = Payment(
        payment_date=None,
        amount=None,
        payer_name="First payer",
        description="First payment",
        reference_number=None,
        source_bank="First bank",
        original_record=None,
    )
    second_payment = Payment(
        payment_date=None,
        amount=None,
        payer_name="Second payer",
        description="Second payment",
        reference_number=None,
        source_bank="Second bank",
        original_record=None,
    )

    class TestImporter(BankImporter):
        def __init__(self, payment: Payment) -> None:
            self.payment = payment

        def import_statement(self, path: Path) -> list[Payment]:
            return [self.payment]

    service = ImportService()
    importers = iter((TestImporter(first_payment), TestImporter(second_payment)))
    service._detect_bank = lambda path: next(importers)

    result = service.import_payments(
        workbook_path,
        (first_statement_path, second_statement_path),
    )

    assert result.payments == [first_payment, second_payment]


def test_import_payments_delegates_matching_to_member_matcher(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "membership.xlsx"
    statement_path = tmp_path / "statement.csv"

    workbook = Workbook()
    workbook.save(workbook_path)
    statement_path.write_bytes(b"statement")

    payment = Payment(
        payment_date=None,
        amount=None,
        payer_name="Payer",
        description="Payment",
        reference_number=None,
        source_bank="Bank",
        original_record=None,
    )

    class TestImporter(BankImporter):
        def import_statement(self, path: Path) -> list[Payment]:
            return [payment]

    class TestMemberMatcher:
        def __init__(self) -> None:
            self.payments: list[Payment] = []

        def match(self, matched_payment: Payment) -> None:
            self.payments.append(matched_payment)

    service = ImportService()
    member_matcher = TestMemberMatcher()
    service._member_matcher = member_matcher
    service._detect_bank = lambda path: TestImporter()

    result = service.import_payments(workbook_path, (statement_path,))

    assert result.payments == [payment]
    assert member_matcher.payments == [payment]
