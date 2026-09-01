from openpyxl import Workbook

from membership_importer.excel.workbook_analyzer import WorkbookAnalysis, WorkbookAnalyzer


def test_analyzer_returns_workbook_structure() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Memberships 2026"
    worksheet.append(["Report", None, None, None])
    worksheet.append(["MAC", "Full name", "January", "Feb"])
    worksheet.append(["MAC-001", "First Member", 10, 20])
    worksheet.append(["MAC-002", "Second Member", 5, 0])

    analysis = WorkbookAnalyzer().analyze(workbook)

    assert isinstance(analysis, WorkbookAnalysis)
    assert analysis.worksheet_names == ("Memberships 2026",)
    assert analysis.active_worksheet == "Memberships 2026"
    assert analysis.header_row == 2
    assert analysis.first_data_row == 3
    assert analysis.last_data_row == 4
    assert analysis.detected_year == "2026"
    assert analysis.month_columns == {"january": 3, "february": 4}
    assert analysis.mac_column == 1


def test_analyzer_does_not_modify_workbook() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "2025"
    worksheet.append(["MAC", "January"])
    worksheet.append(["MAC-001", 10])
    before = [tuple(cell.value for cell in row) for row in worksheet.iter_rows()]

    WorkbookAnalyzer().analyze(workbook)

    after = [tuple(cell.value for cell in row) for row in worksheet.iter_rows()]
    assert after == before
